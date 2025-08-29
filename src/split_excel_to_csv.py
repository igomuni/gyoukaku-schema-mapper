import os
import glob
import csv
from openpyxl import load_workbook

def clean_value(value):
    if value is None: return ""
    return str(value).replace('\n', ' ').replace('\r', ' ')

def excel_to_full_csv(file_path, output_dir):
    print(f"\n--- Excel -> CSV 変換開始: {os.path.basename(file_path)} ---")
    try:
        workbook = load_workbook(filename=file_path, read_only=True)
        for sheet_name in workbook.sheetnames:
            print(f"  シート '{sheet_name}' を処理中...")
            sheet = workbook[sheet_name]
            base_name = os.path.splitext(os.path.basename(file_path))[0]
            safe_sheet_name = "".join(c if c.isalnum() else '_' for c in sheet_name)
            output_filename = f"{base_name}_{safe_sheet_name}.csv"
            output_file_path = os.path.join(output_dir, output_filename)
            try:
                with open(output_file_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
                    writer = csv.writer(csvfile)
                    for row_cells in sheet.iter_rows():
                        current_row_values = [clean_value(cell.value) for cell in row_cells]
                        writer.writerow(current_row_values)
                print(f"  -> ファイルを作成しました: {output_filename}")
            except Exception as e:
                print(f"  -> エラー: CSV書き込み中に問題が発生: {e}")
    except Exception as e:
        print(f"  エラー: Excel読み込み中に問題が発生: {e}")

if __name__ == "__main__":
    print("★★★ ステップ1: ExcelからCSVへの変換処理を開始します ★★★")
    project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    excel_dir = os.path.join(project_root, 'data', 'excel')
    csv_dir = os.path.join(project_root, 'data', 'csv')
    os.makedirs(csv_dir, exist_ok=True)

    excel_files = glob.glob(os.path.join(excel_dir, '*.xlsx'))
    if not excel_files:
        print(f"\n[エラー] '{excel_dir}' 内に処理対象のExcelファイル (.xlsx) が見つかりません。")
    else:
        print(f"\n対象ファイル: {[os.path.basename(f) for f in excel_files]}")
        for file_path in excel_files:
            excel_to_full_csv(file_path, csv_dir)
        print("\n★★★ ステップ1が完了しました ★★★")